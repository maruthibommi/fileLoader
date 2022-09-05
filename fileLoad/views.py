from calendar import c
from django.shortcuts import render
from django.http import HttpResponse
from django.shortcuts import render
from django.core.files.storage import FileSystemStorage
from django.core.files.storage import Storage
import pandas as pd
import numpy as np
import json
from io import BytesIO
import openpyxl
from django.db import connection

def read_file(file_name):
    fs = FileSystemStorage()
    uploaded_file = fs.open(file_name)
    df = openpyxl.load_workbook(uploaded_file,data_only=True)
    for i in df.sheetnames:
        excel_data = list()
        for row in df[i].iter_rows():
            row_data = list()
            for cell in row:
                row_data.append(cell.value)
            excel_data.append(row_data)
    return excel_data

def connect_sql():
    with connection.cursor() as cursor:
        cursor.execute("select * from dbo.[LD.Load_details];")
        df = cursor.fetchall() 
    return df

def upload(request):
    context = {}
    sqlTable = connect_sql()
    validation_table = np.asarray([[k.rstrip() for k in sublist ] for sublist in sqlTable])
    context['fileList'] = validation_table[:,0]
    print(context['fileList'])
    if request.method == 'POST':
        uploaded_file = request.FILES['documents']
        fs = FileSystemStorage()
        if(fs.exists(uploaded_file.name)):
            fs.delete(uploaded_file.name)
        name = fs.save(uploaded_file.name,uploaded_file)
        context['name'] = uploaded_file.name
        context['url'] = fs.url(name)
        context['excel_data'] = read_file(context['name'])
        print(request.POST['fileNames'])
    return render(request,'fileLoad/Html_css_files/mainLoad.html',context)


def validate(request):
    context = {}
    if(request.method == 'POST'):
        context['name'] = request.POST['file_name']
        fs = FileSystemStorage()
        context['excel_data'] = read_file(context['name'])
        data_frame = convert_to_dataframe(context['name'])
        sqlTable = connect_sql()
        validation_table = np.asarray([[k.rstrip() for k in sublist ] for sublist in sqlTable])
        context = data_quality(context['name'],data_frame,validation_table,context) 
        return render(request,'fileLoad/Html_css_files/validate.html',context)

def data_quality(file_name,data_frame,validation_table,context):
    if(file_name == 'data1.csv'):
        df = data_frame[data_frame.duplicated()]
        print(validation_table)
        context['duplicates'] = df.to_numpy()
        return context
    else:
        return context

def convert_to_dataframe(data_path):
    nm_arr = read_file(data_path)
    df = pd.DataFrame(nm_arr)
    return df
    